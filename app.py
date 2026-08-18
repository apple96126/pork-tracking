import streamlit as st
import io
from datetime import date
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import encode_rfc2231

def send_email(to_email, subject, body, attachment=None):
    from_email = "你的gmail帳號@gmail.com"
    password = "你的應用程式密碼"

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email
    msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.getvalue())
        encoders.encode_base64(part)
        filename = "肉品追溯表.xlsx"
        part.add_header(
            "Content-Disposition",
            "attachment",
            filename=encode_rfc2231(filename, "utf-8")
        )
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(from_email, password)
        server.sendmail(from_email, [to_email], msg.as_string())

st.set_page_config(page_title="肉品追溯系統", layout="centered")
st.title("🐖 豬肉追溯系統")

# 輔助函數
def style_range(ws, cell_range, font=None, alignment=None, fill=None, border=None):
    for row in ws[cell_range]:
        for cell in row:
            if font: cell.font = font
            if alignment: cell.alignment = alignment
            if fill: cell.fill = fill
            if border: cell.border = border

def format_date_slash(dt):
    if isinstance(dt, date):
        return f"{dt.year}/{dt.month}/{dt.day}"
    return str(dt)

# =================【資料庫】=================
PORK_HIERARCHY = {
    "香里食品企業股份有限公司": {
        "ME-320039": {"品名": "1.2後腿絞肉206g", "需要QR": True},
        "ME-320040": {"品名": "預設肉品項目206g", "需要QR": True},
        "ME-320018": {"品名": "豬皮", "需要QR": False},
    }
}

# =================【前端輸入】=================
supplier_list = list(PORK_HIERARCHY.keys())
supplier = st.selectbox("1. 請先選取供應商", supplier_list)
available_skus = PORK_HIERARCHY[supplier]
sku_options = [f"{sku} - {info['品名']}" for sku, info in available_skus.items()]
selected_sku_string = st.selectbox("2. 請選取產品品項", sku_options)
selected_sku_code = selected_sku_string.split(" -")[0]
product_name = available_skus[selected_sku_code]["品名"]
need_qr_flow = available_skus[selected_sku_code].get("需要QR", False)

in_date = st.date_input("4. 進貨日", value=date.today())
make_date = st.date_input("11. 製造日期", value=date.today())
valid_date = st.date_input("12. 有效日期", value=date.today())
email_address = st.text_input("5. 請輸入聯絡信箱")

email_valid = False
if email_address:
    if "@" not in email_address or "." not in email_address.split("@")[-1]:
        st.error("⚠️ 信箱格式不正確")
    else:
        st.success("✅ 信箱格式正確")
        email_valid = True

# =================【照片上傳】=================
uploaded_image_files = st.file_uploader("📸 包裝照片（可多選）", type=["jpg","jpeg","png"], accept_multiple_files=True, key="photo1_multi")
uploaded_receipt_file = st.file_uploader("📸 單據照片", type=["jpg","jpeg","png"], key="photo2")
uploaded_qr_screenshot_1 = None
uploaded_qr_screenshot_2 = None
if need_qr_flow:
    uploaded_qr_screenshot_1 = st.file_uploader("📸 QR Code 屠宰日期", type=["jpg","jpeg","png"], key="qr_snap1")
    uploaded_qr_screenshot_2 = st.file_uploader("📸 QR Code 平台", type=["jpg","jpeg","png"], key="qr_snap2")

# =================【Excel 建立】=================
wb = Workbook()
ws = wb.active
ws.title = f"{supplier}_{selected_sku_code}_{product_name}"

# 區塊二：產品包裝圖示
ws.merge_cells('A4:L4')
ws['A4'] = "產品包裝圖示"
style_range(ws, 'A4:L4', font=Font(size=10, bold=True), alignment=Alignment(horizontal="center"))

uploaded_packs = uploaded_image_files[:3] if uploaded_image_files else []
target_pack_cols = ['A5','E5','I5']
if uploaded_packs:
    for idx, img_file in enumerate(uploaded_packs):
        if idx < len(target_pack_cols):
            # 預覽縮小版
            preview_img = PILImage.open(img_file)
            preview_img.thumbnail((150,200))
            st.image(preview_img, caption=f"包裝照片 {idx+1}")
            # Excel 原始解析度
            pil_img = PILImage.open(img_file)
            img_stream = io.BytesIO()
            pil_img.save(img_stream, format='PNG')
            img_stream.seek(0)
            ws.add_image(OpenpyxlImage(img_stream), target_pack_cols[idx])

# 區塊三：原料肉資料 (QR Code)
if need_qr_flow:
    if uploaded_qr_screenshot_1:
        preview_qr1 = PILImage.open(uploaded_qr_screenshot_1)
        preview_qr1.thumbnail((140,240))
        st.image(preview_qr1, caption="QR Code 屠宰日期")
        pil_qr1 = PILImage.open(uploaded_qr_screenshot_1)
        img_stream_qr1 = io.BytesIO()
        pil_qr1.save(img_stream_qr1, format='PNG')
        img_stream_qr1.seek(0)
        ws.add_image(OpenpyxlImage(img_stream_qr1), 'A8')
    if uploaded_qr_screenshot_2:
        preview_qr2 = PILImage.open(uploaded_qr_screenshot_2)
        preview_qr2.thumbnail((250,240))
        st.image(preview_qr2, caption="QR Code 平台")
        pil_qr2 = PILImage.open(uploaded_qr_screenshot_2)
        img_stream_qr2 = io.BytesIO()
        pil_qr2.save(img_stream_qr2, format='PNG')
        img_stream_qr2.seek(0)
        ws.add_image(OpenpyxlImage(img_stream_qr2), 'C8')

# 區塊五：進貨單據
if uploaded_receipt_file:
    preview_receipt = PILImage.open(uploaded_receipt_file)
    preview_receipt.thumbnail((300,200))
    st.image(preview_receipt, caption="進貨單據")
    pil_receipt = PILImage.open(uploaded_receipt_file)
    img_stream_receipt = io.BytesIO()
    pil_receipt.save(img_stream_receipt, format='PNG')
    img_stream_receipt.seek(0)
    ws.add_image(OpenpyxlImage(img_stream_receipt), 'A13')

# 匯出 Excel
excel_data = io.BytesIO()
wb.save(excel_data)
excel_data.seek(0)
download_filename = f"{supplier}_{selected_sku_code}_{product_name}.xlsx"

if email_valid:
    st.download_button("📥下載 Excel 報表", data=excel_data, file_name=download_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if st.button("寄送 Email"):
        try:
            send_email(email_address, "肉品追溯表通知", f"供應商:{supplier}\n品名:{product_name}\n進貨日:{in_date}", excel_data)
            st.success("✅ 已寄送到您的信箱")
        except Exception as e:
            st.error(f"❌ 寄信失敗: {e}")
else:
    st.error("⚠️ 請確認信箱格式正確，才能下載或寄送 Excel 報表")
